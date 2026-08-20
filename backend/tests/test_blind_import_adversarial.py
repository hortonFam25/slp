"""The blind import, attacked.

`test_blind_import.py` proves the feature works and that the happy path leaks
nothing. This module is its adversary: every test here is an attempt to get a
child's name, birthday or state identifier out through the six MCP tools by
somebody who controls BOTH halves of the conversation -- the spreadsheet that
was uploaded AND the model that is orchestrating the import.

That threat model is the real one. The agent driving these tools is the thing
the whole feature is defending the roster from, so "the agent would not do
that" is not an argument available to any assertion in this file.

The sentinels are the existing module's, plus a few of this one's own, and the
rule is the same: they are nonsense on purpose, so a single occurrence anywhere
in a payload is a leak and cannot be a coincidence.
"""

from __future__ import annotations

import io
import json

import pytest

from test_blind_import import (  # noqa: F401  (fixtures are used by name)
    A_DOB,
    A_FIRST,
    A_LAST,
    A_UIC,
    B_DOB,
    B_FIRST,
    B_LAST,
    C_LAST,
    SCHOOL_NAME,
    TEACHER_NAME,
    assert_no_sentinel,
    blob,
    call,
    sheet_of,
    stage,
)

# Adults who exist only here, so a building or a staff name in these files
# cannot be confused with the ones `world` puts in the database.
OTHER_SCHOOL = "Southgate Middle"
THIRD_SCHOOL = "Eastwood Primary"
OTHER_TEACHER = "Devon Ashcombe"
THIRD_TEACHER = "Priya Raghunathan"

# Repeated into a cell long enough to trip the student schema's own length
# limit, which is how a commit gets to fail with a child's cell in its hands.
OVERLONG = "Zzyzxarium"


@pytest.fixture(scope="module")
def world(client):
    """Two therapists of this module's own, and no other rows at all.

    Deliberately not `test_blind_import.world`: that fixture INSERTS its users,
    its school and its teacher, so importing it would run a second time against
    the same sqlite file, collide on `external_auth_id` and leave two buildings
    with the same name for the other module's lookups to pick between. Nothing
    here needs a school or a teacher to exist -- these tests are about what is
    said, not about what resolves.
    """
    from app.db.database import SessionLocal
    from app.models.user import User

    db = SessionLocal()
    try:

        def user(external_id: str) -> int:
            found = (
                db.query(User)
                .filter(User.external_auth_id == external_id)
                .one_or_none()
            )
            if found is None:
                found = User(
                    external_auth_id=external_id,
                    email=f"{external_id}@example.invalid",
                    display_name=external_id,
                    role="therapist",
                    is_active=True,
                )
                db.add(found)
                db.commit()
            return found.id

        return {
            "owner": user("blind-import-adversary-owner"),
            "stranger": user("blind-import-adversary-stranger"),
        }
    finally:
        db.close()


@pytest.fixture(autouse=True)
def no_students_left_behind(client):
    """Sentinel children must not survive from one test into the next."""
    from app.db.database import SessionLocal
    from app.models.import_batch import ImportRow
    from app.models.student import Student
    from app.models.user_student_access import UserStudentAccess

    def sweep():
        db = SessionLocal()
        try:
            ids = [
                row.id
                for row in db.query(Student)
                .filter(Student.last.in_([A_LAST, B_LAST, C_LAST]))
                .all()
            ]
            if ids:
                db.query(ImportRow).filter(
                    ImportRow.resolved_student_id.in_(ids)
                ).update({"resolved_student_id": None}, synchronize_session=False)
                db.query(UserStudentAccess).filter(
                    UserStudentAccess.student_id.in_(ids)
                ).delete(synchronize_session=False)
                db.query(Student).filter(Student.id.in_(ids)).delete(
                    synchronize_session=False
                )
                db.commit()
        finally:
            db.close()

    sweep()
    yield
    sweep()


def revealed_text(preview: dict) -> str:
    """Everything the preview shows verbatim, as one searchable blob."""
    parts = []
    for sheet in preview["sheets"]:
        for candidate in sheet["headerRowCandidates"]:
            parts.extend(str(v) for v in candidate["headerTexts"].values())
        parts.extend(str(column["header"]) for column in sheet["columns"])
    return " | ".join(parts)


# ---------------------------------------------------------------------------
# 1. header_reveal_rows -- the verbatim reveal in the preview
# ---------------------------------------------------------------------------
def test_a_two_row_headerless_sheet_never_prints_row_one(call, world):
    """Two children and no header row. Row 1 is not a heading, it is a child.

    `_differs_from_the_column` compared EXACT shapes, and two surnames of
    different lengths have different shapes -- so a data row "differed from its
    column" merely by being a different length, which is the one thing a data
    row always is.
    """
    rows = [
        [f"{A_LAST}, {A_FIRST}", SCHOOL_NAME, TEACHER_NAME],
        [f"{B_LAST}, {B_FIRST}", OTHER_SCHOOL, OTHER_TEACHER],
    ]
    batch_id, _ = stage(world["owner"], sheet_of(rows, name="Two"))

    preview = call("get_import_preview", batch_id=batch_id)
    assert_no_sentinel("two-row headerless preview", preview)
    assert preview["sheets"][0]["headerRowCandidates"] == [], revealed_text(preview)
    assert preview["sheets"][0]["suggestedHeaderRow"] is None


def test_a_headerless_sheet_of_varying_columns_never_prints_a_child(call, world):
    """The same attack with split name columns and a caseload across buildings.

    Nothing in this sheet is a heading. Every row is a child, every column
    varies row to row, and every cell is a short capitalised word with no digit
    in it -- which is every "looks like a header" heuristic satisfied at once.
    """
    rows = [
        [A_FIRST, A_LAST, SCHOOL_NAME, TEACHER_NAME],
        [B_FIRST, B_LAST, OTHER_SCHOOL, OTHER_TEACHER],
        ["Zephyrandus", C_LAST, THIRD_SCHOOL, THIRD_TEACHER],
        ["Bartholomew", "Fitzwilliam", SCHOOL_NAME, OTHER_TEACHER],
    ]
    batch_id, _ = stage(world["owner"], sheet_of(rows, name="Split"))

    preview = call("get_import_preview", batch_id=batch_id)
    assert_no_sentinel("split-name headerless preview", preview)
    assert preview["sheets"][0]["headerRowCandidates"] == [], revealed_text(preview)


def test_one_wide_row_at_the_bottom_does_not_unstop_the_header_scan(call, world):
    """A footer wider than the header used to keep the scan running into the data.

    The scan stops at the first qualifying row that spans most of the sheet's
    width. `width` was the widest row anywhere -- so a single six-column footer
    under a four-column table meant the real header row never counted as
    full-width, the scan carried on, and row 2 (a child) was printed as a
    second "header candidate".
    """
    rows = [
        ["Student", "Building", "Teacher", "Grade"],
        [f"{A_LAST}, {A_FIRST}", SCHOOL_NAME, TEACHER_NAME, "K"],
        [f"{B_LAST}, {B_FIRST}", OTHER_SCHOOL, OTHER_TEACHER, "1"],
        [f"{C_LAST}, Zephyrandus", THIRD_SCHOOL, THIRD_TEACHER, "2"],
        ["prepared", "by", "the", "district", "data", "team"],
    ]
    batch_id, _ = stage(world["owner"], sheet_of(rows, name="Footer"))

    preview = call("get_import_preview", batch_id=batch_id)
    assert_no_sentinel("footer-widened preview", preview)
    sheet = preview["sheets"][0]
    assert [c["rowIndex"] for c in sheet["headerRowCandidates"]] == [1], revealed_text(
        preview
    )
    assert sheet["suggestedHeaderRow"] == 1


def test_a_real_header_row_is_still_revealed(call, world):
    """The gates must not be so tight that no header is ever printed.

    Without this the safe answer is "reveal nothing", which is safe and
    useless: an agent that cannot read the headings cannot propose a mapping
    and the therapist is back to pasting the spreadsheet into a chat window.
    """
    rows = [
        ["Student", "DOB", "State ID", "Building", "Teacher", "Grade"],
        [
            f"{A_LAST}, {A_FIRST}",
            A_DOB.isoformat(),
            A_UIC,
            SCHOOL_NAME,
            TEACHER_NAME,
            "2",
        ],
        [
            f"{B_LAST}, {B_FIRST}",
            B_DOB.isoformat(),
            "IMPUICSENTINEL009",
            OTHER_SCHOOL,
            OTHER_TEACHER,
            "3",
        ],
    ]
    batch_id, _ = stage(world["owner"], sheet_of(rows, name="Normal"))

    preview = call("get_import_preview", batch_id=batch_id)
    sheet = preview["sheets"][0]
    assert [c["rowIndex"] for c in sheet["headerRowCandidates"]] == [1]
    assert [column["header"] for column in sheet["columns"]][:3] == [
        "Student",
        "DOB",
        "State ID",
    ]
    assert_no_sentinel("ordinary preview", preview)


def test_a_sideways_sheet_never_prints_its_first_row(call, world):
    """One student per COLUMN, with the field names running down the left.

    Row 1 is "Student" followed by two children. Every cell in it is shaped
    unlike "the column below it" -- because there are no columns below it, only
    a birthday sitting on top of a building -- so the row was unanimously
    nominated as the header and printed, children and all.
    """
    rows = [
        ["Student", f"{A_LAST}, {A_FIRST}", f"{B_LAST}, {B_FIRST}"],
        ["DOB", A_DOB.isoformat(), B_DOB.isoformat()],
        ["Building", SCHOOL_NAME, OTHER_SCHOOL],
        ["Teacher", TEACHER_NAME, OTHER_TEACHER],
    ]
    batch_id, _ = stage(world["owner"], sheet_of(rows, name="Sideways"))

    preview = call("get_import_preview", batch_id=batch_id)
    assert_no_sentinel("sideways preview", preview)
    assert preview["sheets"][0]["headerRowCandidates"] == [], revealed_text(preview)


def test_a_sheet_of_no_two_alike_rows_prints_none_of_them(call, world):
    """A column of three different KINDS of thing is not a column.

    With nothing consistent below it, every row in turn is "shaped unlike the
    rows beneath it", and the last one standing gets printed.
    """
    rows = [
        [f"{A_FIRST} {A_LAST}", SCHOOL_NAME],
        [f"{B_LAST}, {B_FIRST}", "Southgate"],
        [C_LAST, f"{THIRD_SCHOOL} District"],
    ]
    batch_id, _ = stage(world["owner"], sheet_of(rows, name="Jumble"))

    preview = call("get_import_preview", batch_id=batch_id)
    assert_no_sentinel("jumbled preview", preview)
    assert preview["sheets"][0]["headerRowCandidates"] == [], revealed_text(preview)


def test_a_header_indistinguishable_from_its_column_is_not_guessed_at(call, world):
    """The deliberate cost of the shape-CLASS comparison, asserted on purpose.

    "First" over "Anna" and "Last" over "Smith" are the same kind of thing in
    the same size, and no shape test can separate the heading from the child.
    The answer is to print neither: the columns come back identified by letter,
    the agent asks the therapist what they are, and nothing is guessed at. If
    somebody ever "fixes" this to print the header, it prints row 1 of a
    header-less sheet too -- which is a child.
    """
    rows = [
        ["First", "Last"],
        [A_FIRST, A_LAST],
        [B_FIRST, B_LAST],
    ]
    batch_id, _ = stage(world["owner"], sheet_of(rows, name="Ambiguous"))

    preview = call("get_import_preview", batch_id=batch_id)
    sheet = preview["sheets"][0]
    assert sheet["headerRowCandidates"] == [], revealed_text(preview)
    assert all(column["header"] is None for column in sheet["columns"])
    assert sheet["dimensions"] == {"rows": 3, "columns": 2}


def test_an_apostrophised_name_row_is_not_a_header(call, world):
    """`looks_like_label` says yes to "O'Brien-Vance". Something else has to say no."""
    rows = [
        ["O'Brien-Vance", SCHOOL_NAME, TEACHER_NAME],
        ["Nakamura-Whitfield", OTHER_SCHOOL, OTHER_TEACHER],
        ["Okonkwo", THIRD_SCHOOL, THIRD_TEACHER],
    ]
    batch_id, _ = stage(world["owner"], sheet_of(rows, name="Punct"))

    preview = call("get_import_preview", batch_id=batch_id)
    assert preview["sheets"][0]["headerRowCandidates"] == [], revealed_text(preview)
    assert "O'Brien-Vance" not in revealed_text(preview)


# ---------------------------------------------------------------------------
# 2. the mapping lie -- declaring a PII column to be an organisational one
# ---------------------------------------------------------------------------
def _mislabel(call, world, column_values, field, sheet="Lie"):
    """Map a column of `column_values` as `field` and run the whole flow."""
    rows = [["Name", "Surname", "Thing"]]
    for index, value in enumerate(column_values):
        rows.append([f"Child{index}", f"Surname{index}", value])
    batch_id, _ = stage(world["owner"], sheet_of(rows, name=sheet))

    mapped = call(
        "set_import_mapping",
        batch_id=batch_id,
        mapping={
            "sheet": sheet,
            "header_row": 1,
            "data_start_row": 2,
            "columns": {"A": "first_name", "B": "last_name", "C": field},
        },
    )
    return batch_id, mapped, call("validate_import", batch_id=batch_id)


def test_a_birthday_column_called_a_school_is_never_quoted(call, world):
    """The lie the design accepts for names must not extend to birthdays.

    An agent may claim any column is the building. The tool's own description
    promises that a date of birth "stays masked forever, whatever it is mapped
    to" -- so the promise has to be enforced on the VALUE, not on the label the
    caller chose to put on it.
    """
    dobs = [A_DOB.isoformat(), B_DOB.isoformat(), "3/17/2011", "May 9, 2013"]
    _, mapped, outcome = _mislabel(call, world, dobs, "school")

    assert_no_sentinel("dob-as-school mapping", mapped)
    assert_no_sentinel("dob-as-school validation", outcome)
    for spelling in dobs:
        assert spelling not in blob(mapped), spelling
        assert spelling not in blob(outcome), spelling


def test_a_birthday_column_called_an_eligibility_is_never_quoted(call, world):
    """`unknown_eligibility` is a WARNING, so the file still commits.

    That makes it the quietest exfiltration channel in the feature: the agent
    labels the birthday column `eligibility`, every row reports it as unknown
    with the value attached, and the import succeeds anyway so nothing looks
    wrong to the therapist.
    """
    dobs = [A_DOB.isoformat(), B_DOB.isoformat(), "07/14/2010"]
    _, mapped, outcome = _mislabel(call, world, dobs, "eligibility")

    assert_no_sentinel("dob-as-eligibility validation", outcome)
    for spelling in dobs:
        assert spelling not in blob(outcome), spelling
        assert spelling not in blob(mapped), spelling


def test_an_identifier_column_called_a_teacher_is_never_quoted(call, world):
    """A state identifier is digits. No adult is called 4820193746."""
    uics = ["4820193746", "4820193747", "0000112233"]
    _, mapped, outcome = _mislabel(call, world, uics, "teacher")

    for uic in uics:
        assert uic not in blob(mapped), uic
        assert uic not in blob(outcome), uic


def test_a_birthday_column_called_a_grade_is_not_sampled(call, world):
    """grade_level is on the reveal allow-list and takes numbers. Not these."""
    dobs = [A_DOB.isoformat(), B_DOB.isoformat(), "2010-01-31"]
    _, mapped, _ = _mislabel(call, world, dobs, "grade_level")

    assert mapped["revealedColumns"]["C"]["sampleValues"] == [], mapped
    for spelling in dobs:
        assert spelling not in blob(mapped), spelling


def test_a_genuine_grade_column_is_still_sampled(call, world):
    """The gate must not eat the values it exists to show."""
    _, mapped, _ = _mislabel(call, world, ["K", "1", "2", "12"], "grade_level")
    assert set(mapped["revealedColumns"]["C"]["sampleValues"]) == {"K", "1", "2", "12"}


def test_a_genuine_school_column_is_still_quoted(call, world):
    """And an unknown building is still reported by name, which is the point."""
    _, mapped, outcome = _mislabel(
        call, world, ["Nrthgate El", "Nrthgate El", "Bldg 4"], "school"
    )
    assert "Nrthgate El" in blob(mapped)
    assert {entry["value"] for entry in outcome["unresolvedValues"]["school"]} == {
        "Nrthgate El",
        "Bldg 4",
    }


def test_the_whole_lie_through_the_real_sdk_call_path(world):
    """Not the tool function -- the thing a client actually drives.

    Every gate in this feature is inside the tool body or the decorator around
    it, and both are reached through `FastMCP.call_tool`. A leak that only
    appears there is the one that ships.
    """
    import asyncio

    from app.mcp import auth as mcp_auth
    from app.mcp.server import mcp_server
    from test_blind_import import principal_for

    rows = [["Name", "Surname", "Born", "Ident"]]
    for first, last, dob, uic in (
        (A_FIRST, A_LAST, A_DOB.isoformat(), "4820193746"),
        (B_FIRST, B_LAST, B_DOB.isoformat(), "4820193747"),
    ):
        rows.append([first, last, dob, uic])
    batch_id, _ = stage(world["owner"], sheet_of(rows, name="Sdk"))

    async def go():
        mapped = await mcp_server.call_tool(
            "set_import_mapping",
            {
                "batch_id": batch_id,
                "mapping": {
                    "sheet": "Sdk",
                    "header_row": 1,
                    "data_start_row": 2,
                    "columns": {
                        "A": "first_name",
                        "B": "last_name",
                        # Both lies, both allow-listed fields.
                        "C": "school",
                        "D": "teacher",
                    },
                },
            },
        )
        checked = await mcp_server.call_tool("validate_import", {"batch_id": batch_id})
        return blob([mapped, checked])

    token = mcp_auth._CURRENT.set(principal_for(world["owner"]))
    try:
        payload = asyncio.run(go())
    finally:
        mcp_auth._CURRENT.reset(token)

    for secret in (A_DOB.isoformat(), B_DOB.isoformat(), "4820193746", "4820193747"):
        assert secret not in payload, secret


def test_the_unresolved_list_is_bounded(call, world):
    """One validate call must not be able to dump a whole column.

    `unresolvedValues` was an entry per DISTINCT spelling with no cap at all,
    so a name column labelled `school` came back as the entire roster in one
    payload -- deduplicated, sorted and tidy.
    """
    from app.services.blind_import import UNRESOLVED_LIST_LIMIT

    names = [f"Perpetua{index} Vandersloot" for index in range(120)]
    _, _, outcome = _mislabel(call, world, names, "school")

    listed = outcome["unresolvedValues"]["school"]
    assert len(listed) <= UNRESOLVED_LIST_LIMIT, len(listed)
    assert outcome["unresolvedValuesTruncated"]["school"] == 120 - len(listed)
    quoted = sum(1 for name in names if name in blob(outcome))
    assert quoted <= UNRESOLVED_LIST_LIMIT, quoted


# ---------------------------------------------------------------------------
# 3. commit: the failure path
# ---------------------------------------------------------------------------
def _long_name_batch(world):
    """Two good rows and a third the student schema will refuse."""
    rows = [
        ["First", "Last"],
        [A_FIRST, A_LAST],
        [B_FIRST, B_LAST],
        [OVERLONG * 15, C_LAST],
    ]
    return stage(world["owner"], sheet_of(rows, name="Fail"))[0]


def _map_names(call, batch_id):
    return call(
        "set_import_mapping",
        batch_id=batch_id,
        mapping={
            "sheet": "Fail",
            "header_row": 1,
            "data_start_row": 2,
            "columns": {"A": "first_name", "B": "last_name"},
        },
    )


def test_a_commit_that_fails_halfway_leaves_nothing_behind(call, world, monkeypatch):
    """All-or-nothing, including when the compensation itself is inconvenient.

    Two students are created and their staged rows point at them; the third row
    fails. Undoing means deleting rows another table already references, and a
    compensation that trips over its own foreign key leaves exactly the thing
    it exists to prevent: children on a caseload from an import that failed.
    """
    from app.db.database import SessionLocal
    from app.models.student import Student
    from app.models.user_student_access import UserStudentAccess
    from app.services import blind_import

    batch_id = _long_name_batch(world)
    _map_names(call, batch_id)

    # The point of this test is the WRITE loop failing, so validation is not
    # allowed to be the thing that saves us here.
    monkeypatch.setattr(
        blind_import, "MAX_IMPORTABLE_VALUE_LENGTH", 10_000, raising=False
    )

    with pytest.raises(Exception) as raised:
        call("commit_import", batch_id=batch_id, confirm=True)

    db = SessionLocal()
    try:
        survivors = (
            db.query(Student).filter(Student.last.in_([A_LAST, B_LAST, C_LAST])).all()
        )
        assert survivors == [], [(s.id, s.last) for s in survivors]
        assert (
            db.query(UserStudentAccess)
            .filter(UserStudentAccess.granted_by_user_id == world["owner"])
            .count()
            == 0
        )
    finally:
        db.close()

    # And the failure itself says nothing about the children in the file.
    message = str(raised.value)
    assert_no_sentinel("failed commit error", message)
    assert OVERLONG not in message, message
    assert "input_value" not in message, message


def test_a_row_the_student_schema_will_refuse_is_a_validation_issue(call, world):
    """Better still: the commit never starts, and the report carries no value."""
    batch_id = _long_name_batch(world)
    _map_names(call, batch_id)

    outcome = call("validate_import", batch_id=batch_id)
    assert outcome["readyToCommit"] is False
    assert outcome["issueCounts"].get("value_too_long") == 1
    issue = next(i for i in outcome["issues"] if i["issue"] == "value_too_long")
    assert issue["row"] == 4
    assert issue["field"] == "first_name"
    assert "value" not in issue
    assert OVERLONG not in blob(outcome)


# ---------------------------------------------------------------------------
# 4. the upload door
# ---------------------------------------------------------------------------
def test_the_uploaded_filename_never_becomes_mcp_output(client, call, world):
    """A CSV's sheet name was its FILE name, and a file name is a name.

    "Ramirez caseload.csv" names a child as surely as a cell does, and the
    service's own docstring promises the filename is never returned over MCP.
    It was -- as `sheets[0].sheet`, and again in every mapping error that lists
    the file's sheets.
    """
    from app.db.database import SessionLocal
    from app.services import blind_import

    db = SessionLocal()
    try:
        batch, secret = blind_import.create_batch(db, world["owner"])
        batch_id = batch.id
    finally:
        db.close()

    body = f"First,Last\r\n{A_FIRST},{A_LAST}\r\n".encode("utf-8")
    response = client.post(
        f"/import/upload/{secret}",
        files={"file": (f"{C_LAST} caseload.csv", body, "text/csv")},
    )
    assert response.status_code == 200, response.text

    preview = call("get_import_preview", batch_id=batch_id)
    assert C_LAST not in blob(preview), blob(preview)[:800]


def test_the_result_page_does_not_echo_the_filename(client, world):
    """The filename is attacker-controlled text on a page with no escaping budget."""
    from app.db.database import SessionLocal
    from app.services import blind_import

    db = SessionLocal()
    try:
        _, secret = blind_import.create_batch(db, world["owner"])
    finally:
        db.close()

    nasty = "<img src=x onerror=alert(1)>\"'.csv"
    response = client.post(
        f"/import/upload/{secret}",
        files={"file": (nasty, b"First,Last\r\nAda,Lovelace\r\n", "text/csv")},
    )
    assert response.status_code == 200
    assert "<img src=x" not in response.text
    assert "onerror" not in response.text


def test_a_traversing_filename_is_not_a_path(client, world):
    from app.db.database import SessionLocal
    from app.models.import_batch import ImportRow
    from app.services import blind_import

    db = SessionLocal()
    try:
        batch, secret = blind_import.create_batch(db, world["owner"])
        batch_id = batch.id
    finally:
        db.close()

    response = client.post(
        f"/import/upload/{secret}",
        files={
            "file": (
                "../../../../etc/passwd.csv",
                b"First,Last\r\nAda,Lovelace\r\n",
                "text/csv",
            )
        },
    )
    assert response.status_code == 200

    db = SessionLocal()
    try:
        sheets = {
            row.sheet_name
            for row in db.query(ImportRow).filter(ImportRow.batch_id == batch_id).all()
        }
        assert all(".." not in name and "/" not in name for name in sheets), sheets
    finally:
        db.close()


def test_the_row_cap_counts_every_sheet_and_stops_reading_at_it(monkeypatch):
    """10,000 rows in a thousand tabs is still 10,000 rows.

    And the cap has to bite while the file is being READ. Counting the rows
    after building every one of them in memory is a way to be handed a 5 MB
    file that expands into a gigabyte of strings before anybody objects.
    """
    import openpyxl

    from app.services import blind_import
    from app.services.blind_import import MAX_DATA_ROWS, UploadRejected

    workbook = openpyxl.Workbook()
    del workbook[workbook.sheetnames[0]]
    for tab in range(4):
        sheet = workbook.create_sheet(f"Tab{tab}")
        for index in range(1800):
            sheet.append([f"r{index}", f"c{index}"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    seen = {"cells": 0}
    real = blind_import._cell_text

    def counted(value):
        seen["cells"] += 1
        return real(value)

    monkeypatch.setattr(blind_import, "_cell_text", counted)

    with pytest.raises(UploadRejected):
        blind_import.parse_upload("huge.xlsx", buffer.getvalue())
    # Two columns per row, plus whatever it took to notice.
    assert seen["cells"] <= (MAX_DATA_ROWS + 10) * 2, seen["cells"]


def test_an_absurd_cell_is_not_stored_whole():
    """A 5 MB upload can hold a 200 MB cell. It must not become a 200 MB string."""
    import openpyxl

    from app.services.blind_import import MAX_CELL_CHARS, parse_upload

    workbook = openpyxl.Workbook()
    workbook.active.append(["First", "Last"])
    workbook.active.append(["Ada", "z" * 400_000])
    buffer = io.BytesIO()
    workbook.save(buffer)

    sheets = parse_upload("big.xlsx", buffer.getvalue())
    assert len(sheets[0]["rows"][1][1][1]) <= MAX_CELL_CHARS


def test_a_csv_with_an_unterminated_quote_is_a_refusal_not_a_crash(client, world):
    """`csv` gives up past 128 KB in one field, and its message is about the file."""
    from app.db.database import SessionLocal
    from app.services import blind_import

    db = SessionLocal()
    try:
        _, secret = blind_import.create_batch(db, world["owner"])
    finally:
        db.close()

    body = ('First,Last\r\nAda,"' + "z" * 200_000).encode("utf-8")
    response = client.post(
        f"/import/upload/{secret}", files={"file": ("odd.csv", body, "text/csv")}
    )
    assert response.status_code == 400
    assert "quotation mark" in response.text


# ---------------------------------------------------------------------------
# 5. ownership, oracles and the sanitizer floor
# ---------------------------------------------------------------------------
def test_a_foreign_batch_is_indistinguishable_from_one_that_never_existed(call, world):
    """Two different failures must not be two different messages."""
    batch_id, _ = stage(world["owner"], sheet_of([["First", "Last"], ["Ada", "L"]]))

    with pytest.raises(ValueError) as foreign:
        call("get_import_preview", batch_id=batch_id, user_id=world["stranger"])
    with pytest.raises(ValueError) as missing:
        call("get_import_preview", batch_id=999_999, user_id=world["stranger"])

    assert str(foreign.value).replace(str(batch_id), "N") == str(
        missing.value
    ).replace("999999", "N")


def test_a_batch_is_bound_to_its_creator_and_takes_no_argument_saying_otherwise(
    call, world
):
    """`create_import_upload` has no user parameter, and the grant names the caller."""
    import inspect

    from app.db.database import SessionLocal
    from app.models.import_batch import ImportBatch
    from app.mcp.server import registered_tools

    tool = {t.name: t for t in registered_tools()}["create_import_upload"]
    assert list(inspect.signature(tool.fn).parameters) == []

    created = call("create_import_upload", user_id=world["stranger"])

    db = SessionLocal()
    try:
        batch = db.get(ImportBatch, created["batchId"])
        assert batch.user_id == world["stranger"]
    finally:
        db.close()


def test_staged_cells_are_stripped_at_every_depth():
    """The denylist is a floor, so it has to hold wherever a payload puts them."""
    from app.mcp.privacy import sanitize_tool_result

    payload = {
        "sheets": [
            {"rows": [{"cells": [A_FIRST, A_LAST], "rowIndex": 4}]},
            {"nested": {"deeper": {"cellsJson": json.dumps([A_UIC])}}},
        ],
        "raw_row": [A_FIRST],
    }
    cleaned = blob(sanitize_tool_result(payload, ()))
    for key in ("cells", "cellsJson", "raw_row"):
        assert key not in cleaned, cleaned


def test_the_length_table_still_matches_the_student_schema():
    """A limit raised in the schema and not here is a commit that fails mid-loop."""
    from app.schemas.student import StudentCreate
    from app.services.blind_import import MAX_IMPORTABLE_VALUE_LENGTH

    schema_field = {
        "first_name": "first",
        "last_name": "last",
        "uic": "uic",
        "grade_level": "grade_level",
        "enrollment_status": "enrollment_status",
    }
    for field, limit in MAX_IMPORTABLE_VALUE_LENGTH.items():
        info = StudentCreate.model_fields[schema_field[field]]
        declared = [
            getattr(item, "max_length")
            for item in info.metadata
            if getattr(item, "max_length", None) is not None
        ]
        assert declared == [limit], (field, declared, limit)


def test_the_import_tool_descriptions_carry_nothing_sensitive():
    """tools/list is read before any call is made, by anything holding a token."""
    from app.mcp.server import registered_tools

    names = {
        "create_import_upload",
        "get_import_preview",
        "set_import_mapping",
        "validate_import",
        "commit_import",
        "discard_import",
    }
    for tool in registered_tools():
        if tool.name in names:
            assert_no_sentinel(f"{tool.name} description", tool.description)
