"""The blind staged import: the spreadsheet enters, and the model never reads it.

The claim this feature makes is a strong one -- a therapist's roster export,
with real names, real birthdays and real state identifiers, passes through the
server and is turned into caseload records WITHOUT any of those values crossing
the MCP connection. A claim like that is worth nothing asserted; every test
here is written to break it if it stops being true.

The method is the drift suite's (`test_mcp_pii.py`): the spreadsheet is built
in-test out of sentinels that are nonsense on purpose, so any occurrence
anywhere in any payload is a leak and cannot be a coincidence. What is
different here is that the sentinels name children who DO NOT EXIST in the
database yet, which is precisely the case the roster-based scrubber cannot
help with -- there is no roster entry to match them against. Everything that
protects them is the masking layer, and that is what these tests are pointed
at.

Covered: the upload token's whole lifecycle; xlsx and csv parsing including the
awkward shapes real exports have (junk banner rows, a second sheet, a merged
cell, a formula cell); the preview containing no sentinel at all; the
mapping-gated reveal showing allow-listed columns and nothing else; validation
reporting row numbers rather than values; a commit that writes REAL values to
the database and returns only aliases; batch ownership; and a discard that
actually destroys the staged copy.
"""

from __future__ import annotations

import asyncio
import io
import json
from datetime import date, datetime, timedelta

import pytest

# ---------------------------------------------------------------------------
# sentinels -- children who do not exist yet, which is the whole point
# ---------------------------------------------------------------------------
A_FIRST, A_LAST = "Xylophonica", "Brekkenridge"
A_UIC = "IMPUICSENTINEL001"
A_DOB = date(2013, 5, 9)

B_FIRST, B_LAST = "Quorvalindra", "Tchaikovskaya"
B_UIC = "IMPUICSENTINEL002"
B_DOB = date(2012, 11, 2)

C_FIRST, C_LAST = "Zephyrandus", "Wollstonecroft"
C_UIC = "IMPUICSENTINEL003"

# A cell that must never be echoed back: with data_only=True a formula yields
# its cached value, and a file written by openpyxl has none -- so this string
# must simply vanish.
FORMULA = "=CONCATENATE(A4,B4)"

# The banner row of a real district export names the therapist and the school
# and is exactly the sort of thing a naive header heuristic prints.
BANNER = f"Speech caseload for {A_FIRST} {A_LAST} 2026"

SCHOOL_NAME = "Northgate Elementary"
SCHOOL_MISSPELLED = "Nrthgate El"
TEACHER_NAME = "Marla Pennington"

SENTINELS = (
    A_FIRST,
    A_LAST,
    A_UIC,
    B_FIRST,
    B_LAST,
    B_UIC,
    C_FIRST,
    C_LAST,
    C_UIC,
    FORMULA,
    BANNER,
    A_DOB.isoformat(),
    f"{A_DOB.month}/{A_DOB.day}/{A_DOB.year}",
    f"{A_DOB.month:02d}/{A_DOB.day:02d}/{A_DOB.year}",
    B_DOB.isoformat(),
    f"{B_DOB.month:02d}/{B_DOB.day:02d}/{B_DOB.year}",
)


def blob(value) -> str:
    return json.dumps(value, default=str)


def assert_no_sentinel(label: str, value) -> None:
    """The whole assertion, applied to one payload."""
    text = blob(value).lower()
    for sentinel in SENTINELS:
        assert sentinel.lower() not in text, (
            f"{label} leaked {sentinel!r}:\n{blob(value)[:3000]}"
        )


# ---------------------------------------------------------------------------
# the workbook
# ---------------------------------------------------------------------------
def build_workbook() -> bytes:
    """A caseload export shaped like the ones that actually arrive.

    Sheet 1 carries a MERGED banner row (value in the top-left, None in the
    rest of the range), a blank row, the real header row, three students, and a
    FORMULA cell. Sheet 2 is a legend tab, which every district export seems to
    have and which no importer should choke on.
    """
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Caseload"

    sheet["A1"] = BANNER
    sheet.merge_cells("A1:F1")
    # Row 2 left entirely empty.
    for column, heading in zip(
        "ABCDEF", ["Student", "DOB", "State ID", "Building", "Teacher", "Grade"]
    ):
        sheet[f"{column}3"] = heading

    sheet.append([f"{A_LAST}, {A_FIRST}", A_DOB, A_UIC, SCHOOL_NAME, TEACHER_NAME, "2"])
    sheet.append([f"{B_LAST}, {B_FIRST}", B_DOB, B_UIC, SCHOOL_NAME, TEACHER_NAME, "3"])
    # Row 6: an unknown building, an unparseable birthday, and a formula in the
    # grade cell.
    sheet.append(
        [f"{C_LAST}, {C_FIRST}", "not a date", C_UIC, SCHOOL_MISSPELLED, TEACHER_NAME, None]
    )
    sheet["F6"] = FORMULA

    legend = workbook.create_sheet("Legend")
    legend.append(["Code", "Meaning"])
    legend.append(["SLI", "Speech Language Impairment"])
    legend.append(["OHI", "Other Health Impairment"])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_csv() -> bytes:
    return (
        "Student,DOB,State ID,Building,Teacher,Grade\r\n"
        f'"{A_LAST}, {A_FIRST}",{A_DOB.isoformat()},{A_UIC},{SCHOOL_NAME},{TEACHER_NAME},2\r\n'
        f'"{B_LAST}, {B_FIRST}",{B_DOB.isoformat()},{B_UIC},{SCHOOL_NAME},{TEACHER_NAME},3\r\n'
    ).encode("utf-8")


MAPPING = {
    "sheet": "Caseload",
    "header_row": 3,
    "data_start_row": 4,
    "columns": {
        "A": "full_name_last_first",
        "B": "date_of_birth",
        "C": "uic",
        "D": "school",
        "E": "teacher",
        "F": "grade_level",
    },
}


# ---------------------------------------------------------------------------
# fixtures
# ---------------------------------------------------------------------------
@pytest.fixture(scope="module")
def world(client):
    """A user to own batches, a school and a teacher for rows to resolve against.

    Depends on ``client`` only for its side effect: the app's startup handler is
    what runs ``create_all`` against the throwaway sqlite file.
    """
    from app.db.database import SessionLocal
    from app.models.school import School
    from app.models.teacher import Teacher
    from app.models.user import User

    db = SessionLocal()
    try:
        owner = User(
            external_auth_id="blind-import-owner",
            email="owner@example.invalid",
            display_name="Import Owner",
            role="therapist",
            is_active=True,
        )
        stranger = User(
            external_auth_id="blind-import-stranger",
            email="stranger@example.invalid",
            display_name="Somebody Else",
            role="therapist",
            is_active=True,
        )
        school = School(name=SCHOOL_NAME, district="Northgate ISD", is_active=True)
        teacher = Teacher(
            first_name="Marla", last_name="Pennington", title="Teacher", is_active=True
        )
        db.add_all([owner, stranger, school, teacher])
        db.commit()
        return {
            "owner": owner.id,
            "stranger": stranger.id,
            "school": school.id,
            "teacher": teacher.id,
        }
    finally:
        db.close()


def principal_for(user_id: int):
    from app.mcp.auth import McpPrincipal

    return McpPrincipal(
        user_id=user_id,
        token_id=1,
        user_name="Pytest Therapist",
        role="therapist",
        is_admin=False,
        access_mode="enforce",
        enforce_access=True,
        allowed_student_ids=[],
    )


@pytest.fixture
def call(world):
    """Invoke a registered MCP tool the way the server does.

    Goes through the live FastMCP registry and the `@tool()` wrapper -- i.e.
    the real sanitizer -- inside the contextvar the auth middleware would have
    set. Nothing here reaches past the door a client comes through.
    """
    from app.mcp import auth as mcp_auth
    from app.mcp.server import registered_tools

    tools = {tool.name: tool for tool in registered_tools()}

    def run(name: str, user_id: int | None = None, **kwargs):
        token = mcp_auth._CURRENT.set(principal_for(user_id or world["owner"]))
        try:
            return tools[name].fn(**kwargs)
        finally:
            mcp_auth._CURRENT.reset(token)

    return run


def stage(user_id: int, sheets, filename: str = "roster.xlsx"):
    """Put rows into a fresh batch directly, for tests that are not about upload."""
    from app.db.database import SessionLocal
    from app.services import blind_import

    db = SessionLocal()
    try:
        batch, secret = blind_import.create_batch(db, user_id)
        batch.filename = filename
        blind_import.store_rows(db, batch, sheets)
        return batch.id, secret
    finally:
        db.close()


def sheet_of(rows, name="Caseload"):
    return [{"name": name, "rows": list(enumerate(rows, start=1))}]


@pytest.fixture(autouse=True)
def no_students_left_over_from_the_last_test(client):
    """Every test starts with the sentinel children NOT yet in the caseload.

    They are the same three children in every workbook this module builds, and
    a commit puts their UICs in the students table -- after which the next
    test's batch validates as `duplicate_uic_existing` and every assertion
    downstream of "ready to commit" quietly stops testing what it says it does.
    Rather than staggering identifiers across tests (which would mean the
    sentinel search has to know about the staggering), the caseload is put back
    the way it was.
    """
    from app.db.database import SessionLocal
    from app.models.import_batch import ImportRow
    from app.models.student import Student
    from app.models.user_student_access import UserStudentAccess

    db = SessionLocal()
    try:
        ids = [
            row.id
            for row in db.query(Student)
            .filter(Student.uic.like("IMPUICSENTINEL%"))
            .all()
        ]
        if ids:
            # The staged rows point at them, and sqlite has FK enforcement on.
            db.query(ImportRow).filter(
                ImportRow.resolved_student_id.in_(ids)
            ).update({"resolved_student_id": None}, synchronize_session=False)
            db.query(UserStudentAccess).filter(
                UserStudentAccess.student_id.in_(ids)
            ).delete(synchronize_session=False)
            db.query(Student).filter(Student.id.in_(ids)).delete(
                synchronize_session=False
            )
            db.commit()
    finally:
        db.close()
    yield


# ---------------------------------------------------------------------------
# 1. the upload token
# ---------------------------------------------------------------------------
def test_a_fresh_token_opens_its_batch(world):
    from app.db.database import SessionLocal
    from app.services import blind_import

    db = SessionLocal()
    try:
        batch, secret = blind_import.create_batch(db, world["owner"])
        assert secret.startswith("slpu_")
        assert len(secret) == len("slpu_") + 40
        # Only the digest is stored -- the secret cannot be recovered.
        assert secret not in (batch.upload_token_hash or "")
        assert blind_import.resolve_upload_batch(db, secret).id == batch.id
    finally:
        db.close()


def test_an_expired_token_is_refused(world):
    from app.db.database import SessionLocal
    from app.services import blind_import

    db = SessionLocal()
    try:
        batch, secret = blind_import.create_batch(db, world["owner"])
        batch.token_expires_at = datetime.utcnow() - timedelta(minutes=1)
        db.commit()
        with pytest.raises(blind_import.UploadRejected) as raised:
            blind_import.resolve_upload_batch(db, secret)
        assert raised.value.status_code == 410
    finally:
        db.close()


def test_a_used_token_is_refused_and_says_so(world):
    """Single use. The 409 rather than a 404 is what tells the therapist why."""
    from app.db.database import SessionLocal
    from app.services import blind_import

    db = SessionLocal()
    try:
        batch, secret = blind_import.create_batch(db, world["owner"])
        blind_import.store_rows(db, batch, sheet_of([["Student"], ["x"]]))
        with pytest.raises(blind_import.UploadRejected) as raised:
            blind_import.resolve_upload_batch(db, secret)
        assert raised.value.status_code == 409
    finally:
        db.close()


def test_a_token_from_one_batch_does_not_open_another(world):
    """The digest is the batch's, not the user's."""
    from app.db.database import SessionLocal
    from app.services import blind_import

    db = SessionLocal()
    try:
        one, secret_one = blind_import.create_batch(db, world["owner"])
        two, _ = blind_import.create_batch(db, world["owner"])
        assert blind_import.resolve_upload_batch(db, secret_one).id == one.id
        assert blind_import.resolve_upload_batch(db, secret_one).id != two.id
    finally:
        db.close()


def test_a_forged_token_is_refused(world):
    from app.db.database import SessionLocal
    from app.services import blind_import

    db = SessionLocal()
    try:
        for bogus in ("slpu_" + "0" * 40, "slp_abcdef", "", "../../etc/passwd"):
            with pytest.raises(blind_import.UploadRejected):
                blind_import.resolve_upload_batch(db, bogus)
    finally:
        db.close()


# ---------------------------------------------------------------------------
# 2. the upload route
# ---------------------------------------------------------------------------
def test_the_upload_page_renders_and_takes_a_workbook(client, world):
    from app.db.database import SessionLocal
    from app.models.import_batch import STATUS_UPLOADED
    from app.services import blind_import

    db = SessionLocal()
    try:
        batch, secret = blind_import.create_batch(db, world["owner"])
        batch_id = batch.id
    finally:
        db.close()

    page = client.get(f"/import/upload/{secret}")
    assert page.status_code == 200
    assert "Upload your caseload" in page.text
    # Self-contained: nothing is fetched from anywhere else.
    assert "http://" not in page.text and "https://" not in page.text
    assert page.headers["cache-control"] == "no-store"

    posted = client.post(
        f"/import/upload/{secret}",
        files={
            "file": (
                "caseload.xlsx",
                build_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert posted.status_code == 200, posted.text
    assert "Uploaded" in posted.text
    assert_no_sentinel("upload result page", posted.text)

    # A second POST on the same link is refused as already used, not as unknown.
    again = client.post(
        f"/import/upload/{secret}",
        files={"file": ("caseload.xlsx", build_workbook(), "application/octet-stream")},
    )
    assert again.status_code == 409
    assert "already been used" in again.text

    db = SessionLocal()
    try:
        from app.models.import_batch import ImportBatch, ImportRow

        reloaded = db.get(ImportBatch, batch_id)
        assert reloaded.status == STATUS_UPLOADED
        assert reloaded.sheet_count == 2
        # Recorded for the app, never returned over MCP.
        assert reloaded.filename == "caseload.xlsx"
        rows = db.query(ImportRow).filter(ImportRow.batch_id == batch_id).all()
        assert {row.sheet_name for row in rows} == {"Caseload", "Legend"}
        # Row 2 was blank and is skipped; the numbering still points at the
        # spreadsheet's own rows.
        caseload = sorted(
            row.row_index for row in rows if row.sheet_name == "Caseload"
        )
        assert caseload == [1, 3, 4, 5, 6]

        banner = next(r for r in rows if r.sheet_name == "Caseload" and r.row_index == 1)
        cells = json.loads(banner.cells_json)
        # A merged range: the value sits in the top-left and nowhere else.
        assert cells == [BANNER]

        last = next(r for r in rows if r.sheet_name == "Caseload" and r.row_index == 6)
        formula_cells = json.loads(last.cells_json)
        # data_only=True: the formula TEXT never reaches storage.
        assert FORMULA not in json.dumps(formula_cells)
        assert len(formula_cells) == 5 or formula_cells[5] is None
    finally:
        db.close()


def test_the_upload_route_refuses_a_wrong_file_type(client, world):
    from app.db.database import SessionLocal
    from app.services import blind_import

    db = SessionLocal()
    try:
        _, secret = blind_import.create_batch(db, world["owner"])
    finally:
        db.close()

    response = client.post(
        f"/import/upload/{secret}",
        files={"file": ("roster.pdf", b"%PDF-1.4 nonsense", "application/pdf")},
    )
    assert response.status_code == 400
    assert "not supported" in response.text
    # A rejected parse must NOT burn the link.
    assert client.get(f"/import/upload/{secret}").status_code == 200


def test_the_upload_route_refuses_an_oversized_file(client, world):
    from app.db.database import SessionLocal
    from app.services import blind_import

    db = SessionLocal()
    try:
        _, secret = blind_import.create_batch(db, world["owner"])
    finally:
        db.close()

    payload = b"a,b,c\r\n" * 900_000  # comfortably past 5 MB
    response = client.post(
        f"/import/upload/{secret}", files={"file": ("big.csv", payload, "text/csv")}
    )
    assert response.status_code == 413


def test_the_upload_route_refuses_an_empty_workbook(client, world):
    import openpyxl

    from app.db.database import SessionLocal
    from app.services import blind_import

    db = SessionLocal()
    try:
        _, secret = blind_import.create_batch(db, world["owner"])
    finally:
        db.close()

    workbook = openpyxl.Workbook()
    buffer = io.BytesIO()
    workbook.save(buffer)
    response = client.post(
        f"/import/upload/{secret}",
        files={"file": ("empty.xlsx", buffer.getvalue(), "application/octet-stream")},
    )
    assert response.status_code == 400
    assert "no rows" in response.text


def test_a_dead_link_says_so_on_the_page(client, world):
    from app.db.database import SessionLocal
    from app.services import blind_import

    db = SessionLocal()
    try:
        batch, secret = blind_import.create_batch(db, world["owner"])
        batch.token_expires_at = datetime.utcnow() - timedelta(minutes=1)
        db.commit()
    finally:
        db.close()

    page = client.get(f"/import/upload/{secret}")
    assert page.status_code == 410
    assert "expired" in page.text


def test_csv_uploads_parse_the_same_way(client, world):
    from app.db.database import SessionLocal
    from app.models.import_batch import ImportRow
    from app.services import blind_import

    db = SessionLocal()
    try:
        batch, secret = blind_import.create_batch(db, world["owner"])
        batch_id = batch.id
    finally:
        db.close()

    response = client.post(
        f"/import/upload/{secret}",
        files={"file": ("caseload.csv", build_csv(), "text/csv")},
    )
    assert response.status_code == 200, response.text

    db = SessionLocal()
    try:
        rows = (
            db.query(ImportRow)
            .filter(ImportRow.batch_id == batch_id)
            .order_by(ImportRow.row_index)
            .all()
        )
        assert [row.row_index for row in rows] == [1, 2, 3]
        assert rows[0].sheet_name == "caseload"
        # The quoted "Last, First" cell survived as ONE cell.
        assert json.loads(rows[1].cells_json)[0] == f"{A_LAST}, {A_FIRST}"
    finally:
        db.close()


# ---------------------------------------------------------------------------
# 3. the preview: shapes and nothing else
# ---------------------------------------------------------------------------
@pytest.fixture
def staged(world):
    from app.services.blind_import import parse_upload

    sheets = parse_upload("caseload.xlsx", build_workbook())
    batch_id, _ = stage(world["owner"], sheets)
    return batch_id


def test_the_preview_contains_no_sentinel_anywhere(call, staged):
    preview = call("get_import_preview", batch_id=staged)
    assert_no_sentinel("get_import_preview", preview)


def test_the_preview_describes_columns_by_shape(call, staged):
    preview = call("get_import_preview", batch_id=staged)
    caseload = next(s for s in preview["sheets"] if s["sheet"] == "Caseload")

    assert caseload["dimensions"]["columns"] == 6
    assert caseload["suggestedHeaderRow"] == 3
    assert caseload["suggestedDataStartRow"] == 4

    by_letter = {column["column"]: column for column in caseload["columns"]}
    # The header row IS revealed -- a mapping cannot be made without it.
    assert by_letter["A"]["header"] == "Student"
    assert by_letter["B"]["header"] == "DOB"

    # The values are not. Every shape in the name column says "Capitalised
    # word, comma, Capitalised word" and nothing else -- which is exactly
    # enough to map it as full_name_last_first and no help at all in working
    # out whose name it is.
    name_shapes = {entry["shape"] for entry in by_letter["A"]["topShapes"]}
    assert all(shape.startswith("Xx") and "," in shape for shape in name_shapes), (
        name_shapes
    )
    assert not any(char.isalpha() and char not in "Xx" for shape in name_shapes
                   for char in shape), name_shapes
    # The banner row above the header is preamble, not data, so it is not
    # counted in with the three students.
    assert by_letter["A"]["nonEmpty"] == 3
    assert by_letter["A"]["distinctValues"] == 3
    # A birthday column is unmistakable by shape and unreadable by value.
    assert {"shape": "####-##-##", "count": 2} in by_letter["B"]["topShapes"]


def test_the_banner_row_is_not_printed(call, staged):
    """The merged title cell names the therapist and a child. It stays masked.

    It is a single populated cell, it carries a year, and it does not differ in
    shape from the column below it in the way a heading does -- any one of those
    is enough to keep it out.
    """
    preview = call("get_import_preview", batch_id=staged)
    caseload = next(s for s in preview["sheets"] if s["sheet"] == "Caseload")
    revealed = [candidate["rowIndex"] for candidate in caseload["headerRowCandidates"]]
    assert 1 not in revealed, caseload["headerRowCandidates"]
    assert revealed == [3], revealed


def test_a_headerless_sheet_reveals_nothing_at_all(call, world):
    """The dangerous case: no header row, so row 1 IS a child.

    Every "looks like a heading" heuristic ever written says row 1 of this
    sheet is a header row -- it is mostly words. It must still not be printed,
    because it is shaped exactly like the column beneath it, which is what
    `header_reveal_rows` checks last and precisely for this.
    """
    rows = [
        [f"{A_LAST}, {A_FIRST}", SCHOOL_NAME, TEACHER_NAME, "K"],
        [f"{B_LAST}, {B_FIRST}", SCHOOL_NAME, TEACHER_NAME, "K"],
        [f"{C_LAST}, {C_FIRST}", SCHOOL_NAME, TEACHER_NAME, "K"],
    ]
    batch_id, _ = stage(world["owner"], sheet_of(rows, name="Raw"))

    preview = call("get_import_preview", batch_id=batch_id)
    assert_no_sentinel("headerless preview", preview)
    sheet = preview["sheets"][0]
    assert sheet["headerRowCandidates"] == []
    assert sheet["suggestedHeaderRow"] is None
    assert all(column["header"] is None for column in sheet["columns"])


def test_the_preview_runs_through_the_real_sdk_call_path(world, staged):
    """The same answer through `FastMCP.call_tool`, which is what a client drives."""
    from app.mcp import auth as mcp_auth
    from app.mcp.server import mcp_server

    async def go():
        return await mcp_server.call_tool("get_import_preview", {"batch_id": staged})

    token = mcp_auth._CURRENT.set(principal_for(world["owner"]))
    try:
        result = asyncio.run(go())
    finally:
        mcp_auth._CURRENT.reset(token)

    assert_no_sentinel("sdk call_tool preview", blob(result))


# ---------------------------------------------------------------------------
# 4. the mapping-gated reveal
# ---------------------------------------------------------------------------
def test_the_mapping_reveals_only_the_allowlisted_columns(call, staged):
    from app.mcp.privacy import SAFE_REVEAL_FIELDS

    result = call("set_import_mapping", batch_id=staged, mapping=MAPPING)

    revealed = result["revealedColumns"]
    assert set(revealed) == {"D", "E", "F"}, revealed
    assert all(entry["field"] in SAFE_REVEAL_FIELDS for entry in revealed.values())
    assert SCHOOL_NAME in revealed["D"]["sampleValues"]
    assert TEACHER_NAME in revealed["E"]["sampleValues"]

    # The name, the birthday and the identifier are withheld -- and named as
    # withheld, so the agent knows the columns exist.
    assert result["withheldColumns"] == {
        "A": "full_name_last_first",
        "B": "date_of_birth",
        "C": "uic",
    }
    assert_no_sentinel("set_import_mapping", result)


def test_mislabelling_the_name_column_as_a_school_is_caught_by_validation(
    call, world
):
    """The allow-list is on the FIELD, so what happens if a field is lied about?

    Nothing stops an agent from claiming the surname column is the building,
    and the reveal then honestly shows that column -- pretending otherwise
    would be a security story that is not true. The protection is that the lie
    does not survive contact with the database: every one of those "schools"
    is unknown, the batch is blocked, and nothing can be written on the
    strength of it. A mislabelled column is a dead end, not a back door.

    This is the honest limit of the design and it is asserted rather than
    glossed: the reveal allow-list defends against a MISTAKE, and the
    unknown-value block is what defends against a deliberate one.
    """
    rows = [
        ["Student", "Building"],
        [f"{A_LAST}, {A_FIRST}", SCHOOL_NAME],
        [f"{B_LAST}, {B_FIRST}", SCHOOL_NAME],
    ]
    batch_id, _ = stage(world["owner"], sheet_of(rows, name="Trick"))

    result = call(
        "set_import_mapping",
        batch_id=batch_id,
        mapping={
            "sheet": "Trick",
            "header_row": 1,
            "data_start_row": 2,
            # A is the NAME column, mapped as the school. B is the school
            # column, mapped as the name.
            "columns": {"A": "school", "B": "full_name_first_last"},
        },
    )
    # It IS revealed -- that is honest, and it is why the next lines matter.
    assert A_LAST in blob(result["revealedColumns"])

    outcome = call("validate_import", batch_id=batch_id)
    assert not outcome["readyToCommit"]
    assert outcome["issueCounts"].get("unknown_school") == 2
    unresolved = outcome["unresolvedValues"]["school"]
    assert {entry["value"] for entry in unresolved} == {
        f"{A_LAST}, {A_FIRST}",
        f"{B_LAST}, {B_FIRST}",
    }
    # And a commit is refused outright rather than half-applied.
    assert call("commit_import", batch_id=batch_id, confirm=True)["committed"] is False


def test_a_revealed_value_is_still_roster_scrubbed(call, world):
    """An EXISTING student's name inside a school cell comes out as the alias."""
    from app.db.database import SessionLocal
    from app.models.student import Student

    db = SessionLocal()
    try:
        existing = Student(
            student_alias="pending",
            first="Marmaduke",
            last="Fenwickshire",
            enrollment_status="Active",
        )
        db.add(existing)
        db.flush()
        existing.student_alias = f"student_{existing.id}"
        db.commit()
        alias = existing.student_alias
    finally:
        db.close()

    rows = [
        ["Student", "Building"],
        ["Doe, Jane", "Fenwickshire Annex"],
    ]
    batch_id, _ = stage(world["owner"], sheet_of(rows, name="Scrub"))

    result = call(
        "set_import_mapping",
        batch_id=batch_id,
        mapping={
            "sheet": "Scrub",
            "header_row": 1,
            "data_start_row": 2,
            "columns": {"A": "full_name_last_first", "B": "school"},
        },
    )
    samples = result["revealedColumns"]["B"]["sampleValues"]
    assert "Fenwickshire" not in blob(samples), samples
    assert alias in blob(samples), samples


def test_a_bad_mapping_is_refused(call, staged):
    for bad, fragment in (
        ({"sheet": "Nope", "columns": {"A": "uic"}}, "no sheet called"),
        (
            {"sheet": "Caseload", "columns": {"A": "wingspan"}},
            "not a field this import understands",
        ),
        (
            {"sheet": "Caseload", "columns": {"1": "uic"}},
            "not a spreadsheet column letter",
        ),
        ({"sheet": "Caseload", "columns": {"A": "uic"}}, "A student needs a name"),
        (
            {
                "sheet": "Caseload",
                "columns": {"A": "first_name", "B": "last_name", "C": "first_name"},
            },
            "both mapped to",
        ),
        (
            {
                "sheet": "Caseload",
                "columns": {"A": "full_name_last_first", "B": "first_name", "C": "last_name"},
            },
            "not both",
        ),
    ):
        with pytest.raises(ValueError) as raised:
            call("set_import_mapping", batch_id=staged, mapping=bad)
        assert fragment in str(raised.value), (bad, str(raised.value))


# ---------------------------------------------------------------------------
# 5. validation
# ---------------------------------------------------------------------------
def test_validation_reports_rows_and_shapes_not_values(call, staged):
    call("set_import_mapping", batch_id=staged, mapping=MAPPING)
    outcome = call("validate_import", batch_id=staged)

    assert_no_sentinel("validate_import", outcome)
    assert outcome["rowsChecked"] == 3
    assert not outcome["readyToCommit"]

    kinds = outcome["issueCounts"]
    assert kinds["unknown_school"] == 1
    assert kinds["unparseable_date"] == 1

    date_issue = next(i for i in outcome["issues"] if i["issue"] == "unparseable_date")
    assert date_issue["row"] == 6
    assert date_issue["column"] == "B"
    assert date_issue["shape"] == "xxx x xxxx"
    assert date_issue["blocking"] is False
    assert "value" not in date_issue

    # The school IS named, because reconciling it is the agent's job.
    school_issue = next(i for i in outcome["issues"] if i["issue"] == "unknown_school")
    assert school_issue["value"] == SCHOOL_MISSPELLED
    assert school_issue["closestExistingMatch"] == SCHOOL_NAME
    assert school_issue["blocking"] is True
    assert outcome["unresolvedValues"]["school"][0]["rows"] == [6]


def test_a_duplicate_uic_is_reported_as_a_row_pair(call, world):
    rows = [
        ["Last", "First", "ID"],
        [A_LAST, A_FIRST, A_UIC],
        [B_LAST, B_FIRST, A_UIC],
    ]
    batch_id, _ = stage(world["owner"], sheet_of(rows, name="Dupes"))
    call(
        "set_import_mapping",
        batch_id=batch_id,
        mapping={
            "sheet": "Dupes",
            "header_row": 1,
            "data_start_row": 2,
            "columns": {"A": "last_name", "B": "first_name", "C": "uic"},
        },
    )
    outcome = call("validate_import", batch_id=batch_id)
    assert_no_sentinel("duplicate uic validation", outcome)

    issue = next(i for i in outcome["issues"] if i["issue"] == "duplicate_uic_in_file")
    assert issue["row"] == 3
    assert issue["duplicateOfRow"] == 2
    assert not outcome["readyToCommit"]


def test_a_duplicate_against_an_existing_student_is_reported_as_an_alias(call, world):
    from app.db.database import SessionLocal
    from app.models.student import Student

    db = SessionLocal()
    try:
        existing = Student(
            student_alias="pending",
            first="Preexisting",
            last="Studentperson",
            uic="EXISTINGUIC42",
            enrollment_status="Active",
        )
        db.add(existing)
        db.flush()
        existing.student_alias = f"student_{existing.id}"
        db.commit()
        alias = existing.student_alias
    finally:
        db.close()

    rows = [["Last", "First", "ID"], [A_LAST, A_FIRST, "EXISTINGUIC42"]]
    batch_id, _ = stage(world["owner"], sheet_of(rows, name="Clash"))
    call(
        "set_import_mapping",
        batch_id=batch_id,
        mapping={
            "sheet": "Clash",
            "header_row": 1,
            "data_start_row": 2,
            "columns": {"A": "last_name", "B": "first_name", "C": "uic"},
        },
    )
    outcome = call("validate_import", batch_id=batch_id)

    issue = next(
        i for i in outcome["issues"] if i["issue"] == "duplicate_uic_existing"
    )
    assert issue["existingStudent"] == alias
    assert "Studentperson" not in blob(outcome)
    assert not outcome["readyToCommit"]


def test_value_overrides_clear_an_unknown_school(call, staged):
    call("set_import_mapping", batch_id=staged, mapping=MAPPING)
    assert not call("validate_import", batch_id=staged)["readyToCommit"]

    call(
        "set_import_mapping",
        batch_id=staged,
        mapping={**MAPPING, "value_overrides": {"school": {SCHOOL_MISSPELLED: SCHOOL_NAME}}},
    )
    outcome = call("validate_import", batch_id=staged)
    assert outcome["readyToCommit"], outcome
    assert outcome["status"] == "validated"
    # The unparseable birthday is a warning, not a blocker.
    assert outcome["warnings"] == 1


# ---------------------------------------------------------------------------
# 6. commit
# ---------------------------------------------------------------------------
@pytest.fixture
def ready(call, staged):
    call(
        "set_import_mapping",
        batch_id=staged,
        mapping={**MAPPING, "value_overrides": {"school": {SCHOOL_MISSPELLED: SCHOOL_NAME}}},
    )
    assert call("validate_import", batch_id=staged)["readyToCommit"]
    return staged


def test_commit_without_confirm_writes_nothing(call, ready):
    from app.db.database import SessionLocal
    from app.models.student import Student

    db = SessionLocal()
    try:
        before = db.query(Student).count()
    finally:
        db.close()

    result = call("commit_import", batch_id=ready, confirm=False)
    assert result["committed"] is False
    assert result["wouldCreate"]["students"] == 3
    assert_no_sentinel("commit_import summary", result)

    db = SessionLocal()
    try:
        assert db.query(Student).count() == before
    finally:
        db.close()


def test_commit_writes_real_values_and_returns_only_aliases(call, ready, world):
    from app.db.database import SessionLocal
    from app.models.student import Student
    from app.models.user_student_access import UserStudentAccess

    result = call("commit_import", batch_id=ready, confirm=True)
    assert result["committed"] is True
    assert result["created"] == 3
    assert_no_sentinel("commit_import", result)
    assert all(alias.startswith("student_") for alias in result["students"])

    db = SessionLocal()
    try:
        # The DATABASE has the real thing. That is the point: the model never
        # saw it, and the record is still correct.
        one = db.query(Student).filter(Student.uic == A_UIC).one()
        assert (one.first, one.last) == (A_FIRST, A_LAST)
        assert one.date_of_birth == A_DOB
        assert one.school_id == world["school"]
        assert one.teacher_id == world["teacher"]
        assert one.grade_level == "2"
        assert one.student_alias == f"student_{one.id}"
        assert one.alias in result["students"]

        two = db.query(Student).filter(Student.uic == B_UIC).one()
        assert two.date_of_birth == B_DOB

        # The row whose building was misspelled is linked through the override.
        three = db.query(Student).filter(Student.uic == C_UIC).one()
        assert three.school_id == world["school"]
        # ...and its unparseable birthday was dropped, not guessed at.
        assert three.date_of_birth is None

        granted = {
            row.student_id
            for row in db.query(UserStudentAccess)
            .filter(UserStudentAccess.user_id == world["owner"])
            .all()
        }
        assert {one.id, two.id, three.id} <= granted
    finally:
        db.close()


def test_a_committed_batch_cannot_be_committed_twice(call, ready):
    call("commit_import", batch_id=ready, confirm=True)
    with pytest.raises(ValueError) as raised:
        call("commit_import", batch_id=ready, confirm=True)
    assert "committed" in str(raised.value)


def test_commit_refuses_while_anything_blocks(call, staged):
    call("set_import_mapping", batch_id=staged, mapping=MAPPING)
    result = call("commit_import", batch_id=staged, confirm=True)
    assert result["committed"] is False
    assert result["blockingIssues"] >= 1
    assert_no_sentinel("blocked commit", result)


# ---------------------------------------------------------------------------
# 7. access and disposal
# ---------------------------------------------------------------------------
def test_another_users_batch_is_not_found(call, world, staged):
    for tool_name, kwargs in (
        ("get_import_preview", {}),
        ("set_import_mapping", {"mapping": MAPPING}),
        ("validate_import", {}),
        ("commit_import", {"confirm": False}),
        ("discard_import", {"confirm": False}),
    ):
        with pytest.raises(ValueError) as raised:
            call(tool_name, user_id=world["stranger"], batch_id=staged, **kwargs)
        assert "not found" in str(raised.value), (tool_name, str(raised.value))


def test_discard_destroys_the_staged_rows(call, staged):
    from app.db.database import SessionLocal
    from app.models.import_batch import ImportBatch, ImportRow

    db = SessionLocal()
    try:
        assert db.query(ImportRow).filter(ImportRow.batch_id == staged).count() > 0
    finally:
        db.close()

    dry = call("discard_import", batch_id=staged, confirm=False)
    assert dry["discarded"] is False
    assert dry["wouldDelete"]["stagedRows"] == 8

    done = call("discard_import", batch_id=staged, confirm=True)
    assert done["discarded"] is True
    assert_no_sentinel("discard_import", done)

    db = SessionLocal()
    try:
        assert db.query(ImportRow).filter(ImportRow.batch_id == staged).count() == 0
        assert db.get(ImportBatch, staged) is None
    finally:
        db.close()


def test_the_raw_cells_never_appear_in_any_import_tool_output(call, world, staged):
    """cells_json is the PII, and no payload from any of these tools carries it.

    Both halves are asserted: no sentinel VALUE, and no KEY that could be
    carrying the raw row -- `app.mcp.privacy` drops those structurally, so a
    future refactor that adds one produces an empty field rather than a leak.
    """
    import re

    payloads = [
        call("create_import_upload"),
        call("get_import_preview", batch_id=staged),
        call("set_import_mapping", batch_id=staged, mapping=MAPPING),
        call("validate_import", batch_id=staged),
        call("commit_import", batch_id=staged, confirm=False),
        call("discard_import", batch_id=staged, confirm=False),
    ]

    forbidden = {"cells", "cellsjson", "rawcells", "rawrow", "rowcells", "cellvalues"}
    normalize = re.compile(r"[^a-z0-9]+")

    def keys(value):
        if isinstance(value, dict):
            for key, item in value.items():
                yield normalize.sub("", str(key).lower())
                yield from keys(item)
        elif isinstance(value, list):
            for item in value:
                yield from keys(item)

    for index, payload in enumerate(payloads):
        assert_no_sentinel(f"import tool #{index}", payload)
        leaked = forbidden & set(keys(payload))
        assert not leaked, (index, leaked, blob(payload)[:2000])


def test_the_import_tools_are_all_behind_the_pii_filter():
    """Belt and braces over the registry walk in test_mcp_pii.py."""
    from app.mcp.server import registered_tools

    expected = {
        "create_import_upload",
        "get_import_preview",
        "set_import_mapping",
        "validate_import",
        "commit_import",
        "discard_import",
    }
    by_name = {tool.name: tool for tool in registered_tools()}
    assert expected <= set(by_name), sorted(expected - set(by_name))
    assert all(
        getattr(by_name[name].fn, "__pii_filtered__", False) for name in expected
    )


# ---------------------------------------------------------------------------
# 8. the masking primitives
# ---------------------------------------------------------------------------
@pytest.mark.parametrize(
    "value,shape",
    [
        ("Ramirez", "Xxxxxxx"),
        ("RAMIREZ", "XXXXXXX"),
        ("Ramirez, Sofia", "Xxxxxxx, Xxxxx"),
        ("3/17/2011", "#/##/####"),
        ("2011-03-17", "####-##-##"),
        ("K", "X"),
        ("", None),
        ("   ", None),
        (None, None),
        (4, "#"),
    ],
)
def test_mask_value(value, shape):
    from app.mcp.privacy import mask_value

    assert mask_value(value) == shape


def test_a_long_value_is_truncated_rather_than_shaped_in_full():
    from app.mcp.privacy import SHAPE_MAX_LENGTH, SHAPE_TRUNCATION_MARKER, mask_value

    shape = mask_value("a" * 400)
    assert shape == "x" * SHAPE_MAX_LENGTH + SHAPE_TRUNCATION_MARKER
    # Exactly at the cap is not truncated, so the marker means what it says.
    assert mask_value("a" * SHAPE_MAX_LENGTH) == "x" * SHAPE_MAX_LENGTH


def test_an_accent_does_not_survive_as_a_fingerprint():
    """NFD spellings must not leave a bare combining mark in the shape."""
    from app.mcp.privacy import mask_value

    composed = mask_value("José García")
    decomposed = mask_value("José García")
    assert composed == decomposed == "Xxxx Xxxxxx"


def test_a_shape_cannot_be_turned_back_into_a_value():
    """Two different children of the same name length share one shape."""
    from app.mcp.privacy import mask_value

    assert mask_value("Ramirez") == mask_value("Bennett") == mask_value("Okonkwo")


def test_looks_like_label_rejects_anything_dated_or_identified():
    from app.mcp.privacy import looks_like_label

    assert looks_like_label("First Name")
    assert looks_like_label("DOB")
    assert not looks_like_label("2026 caseload")
    assert not looks_like_label("3/17/11")
    assert not looks_like_label("UICSENTINEL123456")
    assert not looks_like_label("x" * 200)
    assert not looks_like_label("12345")
